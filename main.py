from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementClickInterceptedException
import time, random, os, json
from openpyxl import Workbook, load_workbook

# ===== Configurações de espera =====
WAIT_SHORT = 10        # cliques/cookies
WAIT_LONG  = 60        # carregamentos de páginas/tabelas
POLL_SLEEP = 0.25      # intervalo do polling leve
POST_CLICK_SLEEP = 0.35  # pausa pós-clique para animações

# ===== Arquivos de saída/checkpoint =====
EXCEL_PATH = "cids.xlsx"
EXCEL_SHEET = "dados"
PROGRESS_PATH = "progress.json"

# ---------- Utilitários Excel (openpyxl) ----------
def ensure_workbook():
    """
    Garante que o arquivo EXCEL_PATH exista com a planilha e cabeçalho.
    """
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = EXCEL_SHEET
        ws.append(["categoria_codigo", "categoria_descricao", "cid_codigo", "cid_descricao"])
        wb.save(EXCEL_PATH)
    else:
        wb = load_workbook(EXCEL_PATH)
        if EXCEL_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(EXCEL_SHEET)
            ws.append(["categoria_codigo", "categoria_descricao", "cid_codigo", "cid_descricao"])
            wb.save(EXCEL_PATH)

def load_processed_categories_from_xlsx():
    """
    Lê o Excel e devolve um set com os códigos de categoria já processados (coluna A).
    Ignora vazios.
    """
    ensure_workbook()
    processed = set()
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[EXCEL_SHEET]
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue  # pula cabeçalho
        if not row:
            continue
        cod = (row[0] or "").strip() if isinstance(row[0], str) else (str(row[0]) if row[0] is not None else "")
        if cod:
            processed.add(cod)
    wb.close()
    return processed

def append_rows_xlsx(rows):
    """
    Acrescenta linhas ao Excel incrementalmente.
    rows: lista de listas [categoria_codigo, categoria_descricao, cid_codigo, cid_descricao]
    """
    ensure_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[EXCEL_SHEET]
    for r in rows:
        ws.append(r)
    wb.save(EXCEL_PATH)
    wb.close()

# ---------- Checkpoint ----------
def load_progress():
    if not os.path.exists(PROGRESS_PATH):
        return {"pagina_atual": 1, "proximo_indice_da_pagina": 0}
    with open(PROGRESS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def save_progress(pagina_atual, proximo_indice_da_pagina):
    data = {"pagina_atual": pagina_atual, "proximo_indice_da_pagina": proximo_indice_da_pagina}
    with open(PROGRESS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ---------- Selenium helpers ----------
opts = Options()
opts.add_argument("--headless=new")
opts.add_argument("--disable-notifications")
opts.add_argument("--disable-gpu")
opts.add_argument("--no-sandbox")
opts.add_argument("--window-size=1366,768")

driver = webdriver.Chrome(options=opts)
wait = WebDriverWait(driver, WAIT_LONG)

def switch_into_categorias(driver, wait):
    """
    Garante que estamos no contexto (main ou iframe) onde #tbCategorias existe.
    Depois de chamar, o driver fica dentro do frame correto (se houver).
    """
    def _in_categorias(drv):
        drv.switch_to.default_content()
        if drv.find_elements(By.ID, "tbCategorias"):
            return True
        frames = drv.find_elements(By.TAG_NAME, "iframe")
        for f in frames:
            drv.switch_to.frame(f)
            if drv.find_elements(By.ID, "tbCategorias"):
                return True
            drv.switch_to.default_content()
        return False

    wait.until(_in_categorias)

    driver.switch_to.default_content()
    if driver.find_elements(By.ID, "tbCategorias"):
        return
    for f in driver.find_elements(By.TAG_NAME, "iframe"):
        driver.switch_to.frame(f)
        if driver.find_elements(By.ID, "tbCategorias"):
            return
        driver.switch_to.default_content()

def scroll_center(elem):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)

def safe_click(elem):
    scroll_center(elem)
    try:
        elem.click()
    except Exception:
        driver.execute_script("arguments[0].click();", elem)

def click_and_wait(clickable, locator_to_wait, max_tries=3, base_sleep=POST_CLICK_SLEEP):
    """
    Clica em `clickable` e espera `locator_to_wait` aparecer.
    Tenta com backoff para lidar com latência/overlay.
    """
    for attempt in range(1, max_tries + 1):
        scroll_center(clickable)
        try:
            clickable.click()
        except (ElementClickInterceptedException, StaleElementReferenceException, Exception):
            driver.execute_script("arguments[0].click();", clickable)

        # micro-pausa para permitir render/transição
        time.sleep(base_sleep + (attempt - 1) * 0.25 + random.uniform(0, 0.15))
        try:
            WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located(locator_to_wait))
            return True
        except TimeoutException:
            if attempt == max_tries:
                return False
            time.sleep(0.4 + 0.2 * attempt)
    return False

def click_voltar():
    """Clica no botão Voltar da tela de detalhes (button id=btnVoltarTbListCategorias).
       Se não achar, usa driver.back() como fallback e espera a tabela principal.
    """
    try:
        btn_voltar = WebDriverWait(driver, WAIT_LONG).until(
            EC.element_to_be_clickable((By.ID, "btnVoltarTbListCategorias"))
        )
        ok = click_and_wait(btn_voltar, (By.ID, "tbCategorias"), max_tries=2)
        if not ok:
            driver.back()
            WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located((By.ID, "tbCategorias")))
    except TimeoutException:
        driver.back()
        WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located((By.ID, "tbCategorias")))

def go_next_page() -> bool:
    """
    Vai para a próxima página da tabela de categorias (DataTables).
    Retorna True se conseguiu avançar, False se já está na última.
    """
    switch_into_categorias(driver, wait)

    # 1) Preferência: DataTables (#tbCategorias_next)
    try:
        next_btn = driver.find_element(By.ID, "tbCategorias_next")
        cls = (next_btn.get_attribute("class") or "").lower()
        if "disabled" in cls:
            return False

        def first_row_key():
            try:
                rows = driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr")
                if not rows:
                    return ""
                tds = rows[0].find_elements(By.TAG_NAME, "td")
                if len(tds) < 2:
                    return ""
                return (tds[0].text or "").strip() + "|" + (tds[1].text or "").strip()
            except Exception:
                return ""

        before_key = first_row_key()
        safe_click(next_btn)

        WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located((By.ID, "tbCategorias")))

        for _ in range(60):  # ~15s
            time.sleep(POLL_SLEEP)
            after_key = first_row_key()
            rows_now = driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr")
            if after_key and after_key != before_key:
                return True
            if rows_now:
                rows_prev = driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr")
                if len(rows_now) != len(rows_prev):
                    return True
        return False
    except Exception:
        pass

    # 2) Fallback: seletores genéricos
    candidatos = [
        (By.XPATH, "//a[contains(., 'Próxima') or contains(., 'Proxima') or contains(., 'Next')]"),
        (By.XPATH, "//button[contains(., 'Próxima') or contains(., 'Proxima') or contains(., 'Next')]"),
        (By.CSS_SELECTOR, "[aria-label='Próxima página'], [aria-label='Proxima página'], [aria-label='Next']"),
        (By.CSS_SELECTOR, ".paginate_button.next, .pagination .next a, .pagination li.next a"),
        (By.XPATH, "//a[.//svg or .//i][contains(@class,'next') or contains(@aria-label,'Próxima') or contains(@aria-label,'Next')]"),
    ]

    def first_row_key():
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr")
            if not rows:
                return ""
            tds = rows[0].find_elements(By.TAG_NAME, "td")
            if len(tds) < 2:
                return ""
            return (tds[0].text or "").strip() + "|" + (tds[1].text or "").strip()
        except Exception:
            return ""

    for by, sel in candidatos:
        try:
            buttons = driver.find_elements(by, sel)
            buttons = [b for b in buttons if b.is_displayed() and b.is_enabled()]
            if not buttons:
                continue

            for btn in buttons:
                cls = (btn.get_attribute("class") or "").lower()
                aria_disabled = (btn.get_attribute("aria-disabled") or "").lower()
                if "disabled" in cls or aria_disabled == "true":
                    continue

                before_key = first_row_key()
                safe_click(btn)

                try:
                    WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located((By.ID, "tbCategorias")))
                    for _ in range(60):  # ~15s
                        time.sleep(POLL_SLEEP)
                        after_key = first_row_key()
                        if after_key and after_key != before_key:
                            return True
                    return False
                except TimeoutException:
                    switch_into_categorias(driver, wait)
                    return True
        except Exception:
            continue
    return False

def set_page_size_100():
    """Seleciona 100 resultados por página no seletor #tbCategorias_length > label > select e
    espera a tabela redesenhar (mais linhas na página)."""
    switch_into_categorias(driver, wait)

    select_el = WebDriverWait(driver, WAIT_LONG).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#tbCategorias_length > label > select"))
    )

    linhas_antes = driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr")
    num_antes = len(linhas_antes)

    try:
        sel = Select(select_el)
        try:
            sel.select_by_value("100")
        except Exception:
            sel.select_by_visible_text("100")
    except Exception:
        driver.execute_script("""
            const s = document.querySelector('#tbCategorias_length > label > select');
            if (s) { s.value = '100'; s.dispatchEvent(new Event('change', {bubbles: true})); }
        """)

    for _ in range(60):  # ~15s
        time.sleep(POLL_SLEEP)
        linhas_depois = driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr")
        if len(linhas_depois) > num_antes or linhas_depois != linhas_antes:
            break

# ========= INÍCIO =========
try:
    # Progresso + categorias já processadas (para evitar duplicados ao retomar)
    progress = load_progress()
    processed_codes = load_processed_categories_from_xlsx()
    pagina_alvo = progress.get("pagina_atual", 1)
    i_alvo = progress.get("proximo_indice_da_pagina", 0)

    driver.get("https://www.cremesp.org.br/?siteAcao=cid10")

    # Aceitar cookies se aparecer
    try:
        btn_cookie = WebDriverWait(driver, WAIT_SHORT).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Ciente') or contains(., 'OK')]"))
        )
        safe_click(btn_cookie)
    except Exception:
        pass

    # Entrar no contexto correto e setar 100 por página
    switch_into_categorias(driver, wait)
    set_page_size_100()

    # ===== LOOP PRINCIPAL: percorre todas as páginas =====
    pagina = 1
    i = 0  # índice corrente na página

    # Retomada: avança até a página alvo
    while pagina < pagina_alvo:
        if not go_next_page():
            break
        pagina += 1
        switch_into_categorias(driver, wait)
        WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located((By.ID, "tbCategorias")))
        set_page_size_100()
        for _ in range(40):
            time.sleep(POLL_SLEEP)
            if driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr"):
                break

    # começa do índice salvo, se houver retomada
    i = i_alvo
    save_progress(pagina, i)

    while True:
        switch_into_categorias(driver, wait)
        linhas = driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr")

        # esgotou as linhas da página? tenta a próxima
        if i >= len(linhas):
            save_progress(pagina + 1, 0)
            if go_next_page():
                pagina += 1
                switch_into_categorias(driver, wait)
                WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located((By.ID, "tbCategorias")))
                set_page_size_100()
                for _ in range(40):
                    time.sleep(POLL_SLEEP)
                    if driver.find_elements(By.CSS_SELECTOR, "#tbCategorias > tbody > tr"):
                        break
                i = 0
                save_progress(pagina, i)
                continue
            else:
                break  # acabou TODAS as páginas

        # ===== processa a linha i desta página =====
        linha = linhas[i]
        tds = linha.find_elements(By.TAG_NAME, "td")

        # precisa ter pelo menos 3 colunas (código, descrição, botão)
        if len(tds) < 3:
            i += 1
            save_progress(pagina, i)
            continue

        codigo = (tds[0].text or "").strip()
        descricao = (tds[1].text or "").strip()

        # pula linhas vazias/placeholder
        if not codigo and not descricao:
            i += 1
            save_progress(pagina, i)
            continue

        print(f"\nCategoria: {codigo} - {descricao}")

        # evita duplicado (já processados e com código não vazio)
        if codigo and codigo in processed_codes:
            print("   (já processada; pulando)")
            i += 1
            save_progress(pagina, i)
            continue

        # botão do olho
        candid = linha.find_elements(By.CSS_SELECTOR, "td:nth-child(3) button")
        if not candid:
            candid = linha.find_elements(By.XPATH, ".//td[3]//button | .//td[3]//a")
        if not candid:
            i += 1
            save_progress(pagina, i)
            continue
        botao = candid[0]

        # abre detalhe (espera pelo botão Voltar)
        ok = click_and_wait(botao, (By.ID, "btnVoltarTbListCategorias"), max_tries=3)
        if not ok:
            driver.execute_script("arguments[0].click();", botao)
            WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located((By.ID, "btnVoltarTbListCategorias")))

        # coleta linhas de CIDs (se houver)
        tabela = driver.find_elements(By.ID, "tabela_body")
        if not tabela:
            tabela = driver.find_elements(By.CSS_SELECTOR, "[id*='tabela_body']")

        detalhas = []
        if tabela:
            for _ in range(int(8 / POLL_SLEEP)):  # ~8s
                detalhas = driver.find_elements(By.CSS_SELECTOR, "[id*='tabela_body'] > tr")
                if detalhas:
                    break
                time.sleep(POLL_SLEEP)

        out_rows = []
        if detalhas:
            for row in detalhas:
                cols = row.find_elements(By.TAG_NAME, "td")
                if len(cols) >= 2:
                    cid_codigo = (cols[0].text or "").strip()
                    cid_desc   = (cols[1].text or "").strip()
                    out_rows.append([codigo, descricao, cid_codigo, cid_desc])
        else:
            # registre categorias sem detalhe, se quiser manter
            out_rows.append([codigo, descricao, "", ""])

        # salva Excel incremental
        append_rows_xlsx(out_rows)
        if codigo:
            processed_codes.add(codigo)

        # volta pra lista
        click_voltar()
        switch_into_categorias(driver, wait)

        # avança pro próximo item da MESMA página
        i += 1
        save_progress(pagina, i)

        time.sleep(0.3)

finally:
    driver.quit()
